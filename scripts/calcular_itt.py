"""
calcular_itt.py
===============
Script de cálculo del Índice de Transformación Territorial (ITT)
para el área del Pulmón de Oriente – Cali 2024–2027.

Flujo:
  1. Lee los indicadores crudos desde `data/excel/indicadores_itt.xlsx`
     (o desde un dict en memoria si se llama como módulo).
  2. Normaliza cada indicador con Min-Max según los umbrales definidos.
  3. Agrega por dimensión (promedio simple de indicadores normalizados).
  4. Calcula el ITT global como suma ponderada de las 5 dimensiones.
  5. Escribe el resultado en `data/itt_pulmon.json`.

Uso:
  python scripts/calcular_itt.py
  python scripts/calcular_itt.py --input data/excel/indicadores_itt.xlsx
  python scripts/calcular_itt.py --input data/excel/indicadores_itt.xlsx --periodo 2026-T1
"""

import json
import argparse
from pathlib import Path
from datetime import datetime

# ── OPENPYXL es la única dependencia extra ──────────────────────────────────
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE DIMENSIONES, PESOS E INDICADORES
# ═══════════════════════════════════════════════════════════════════════════

DIMENSIONES = {
    "seguridad": {
        "nombre": "Seguridad",
        "icono":  "shield",
        "peso":   0.20,
        "color":  "#C1272D",
        "indicadores": [
            {
                "id":       "homicidios",
                "nombre":   "Tasa de homicidios (x 100k hab.)",
                "unidad":   "tasa",
                "fuente":   "Policía Nacional / SIEDCO",
                "oficial":  True,
                "inverso":  True,   # menor = mejor
                "ref_min":  15.0,   # meta 2027 (mejor escenario)
                "ref_max":  45.0,   # línea base 2024-T4
            },
            {
                "id":       "hurtos",
                "nombre":   "Hurtos reportados",
                "unidad":   "casos",
                "fuente":   "SIEDCO",
                "oficial":  True,
                "inverso":  True,
                "ref_min":  150,
                "ref_max":  500,
            },
            {
                "id":       "percepcion_seguridad",
                "nombre":   "Percepción de seguridad",
                "unidad":   "%",
                "fuente":   "Encuesta ITT (estimado)",
                "oficial":  False,
                "inverso":  False,  # mayor = mejor
                "ref_min":  20.0,
                "ref_max":  75.0,
            },
            {
                "id":       "presencia_institucional",
                "nombre":   "Presencia institucional",
                "unidad":   "CAI activos",
                "fuente":   "Secretaría de Seguridad",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  1,
                "ref_max":  5,
            },
        ],
    },
    "entorno_urbano": {
        "nombre": "Entorno Urbano",
        "icono":  "city",
        "peso":   0.30,
        "color":  "#1565C0",
        "indicadores": [
            {
                "id":       "espacio_publico",
                "nombre":   "Espacio público recuperado",
                "unidad":   "ha",
                "fuente":   "Secretaría de Vivienda",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0.0,
                "ref_max":  10.0,
            },
            {
                "id":       "frentes_obra",
                "nombre":   "Obras de infraestructura activas",
                "unidad":   "frentes",
                "fuente":   "UP Secretarías",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  60,
            },
            {
                "id":       "luminarias",
                "nombre":   "Luminarias instaladas",
                "unidad":   "unidades",
                "fuente":   "Emcali / Alumbrado",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  400,
            },
            {
                "id":       "parques",
                "nombre":   "Parques y zonas verdes intervenidos",
                "unidad":   "equipamientos",
                "fuente":   "Secretaría de Deporte",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  15,
            },
        ],
    },
    "movilidad": {
        "nombre": "Movilidad",
        "icono":  "route",
        "peso":   0.15,
        "color":  "#00796B",
        "indicadores": [
            {
                "id":       "tramos_viales",
                "nombre":   "Tramos viales mejorados",
                "unidad":   "km",
                "fuente":   "Infraestructura Vial",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0.0,
                "ref_max":  8.0,
            },
            {
                "id":       "accesibilidad_mio",
                "nombre":   "Accesibilidad a MIO (500m)",
                "unidad":   "%",
                "fuente":   "MIO / Metrocali",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  60.0,
                "ref_max":  90.0,
            },
            {
                "id":       "tiempo_desplazamiento",
                "nombre":   "Tiempo promedio de desplazamiento",
                "unidad":   "min",
                "fuente":   "Encuesta estimada",
                "oficial":  False,
                "inverso":  True,
                "ref_min":  25.0,
                "ref_max":  55.0,
            },
            {
                "id":       "ciclorutas",
                "nombre":   "Ciclorutas activas en el área",
                "unidad":   "km",
                "fuente":   "Secretaría de Movilidad",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0.0,
                "ref_max":  5.0,
            },
        ],
    },
    "desarrollo_social": {
        "nombre": "Desarrollo Social",
        "icono":  "people",
        "peso":   0.20,
        "color":  "#6A1B9A",
        "indicadores": [
            {
                "id":       "ninos_programas",
                "nombre":   "Niños en programas de atención",
                "unidad":   "beneficiarios",
                "fuente":   "Secretaría de Bienestar",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  4000,
            },
            {
                "id":       "cobertura_educativa",
                "nombre":   "Cobertura educativa (preescolar-media)",
                "unidad":   "%",
                "fuente":   "Secretaría de Educación",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  70.0,
                "ref_max":  100.0,
            },
            {
                "id":       "acceso_salud",
                "nombre":   "Acceso a servicios de salud",
                "unidad":   "%",
                "fuente":   "Secretaría de Salud",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  50.0,
                "ref_max":  100.0,
            },
            {
                "id":       "familias_subsidio",
                "nombre":   "Familias en programas de subsidio",
                "unidad":   "familias",
                "fuente":   "Secretaría de Bienestar",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  1200,
            },
        ],
    },
    "actividad_economica": {
        "nombre": "Actividad Económica",
        "icono":  "store",
        "peso":   0.15,
        "color":  "#E65100",
        "indicadores": [
            {
                "id":       "negocios_formalizados",
                "nombre":   "Negocios formalizados en el área",
                "unidad":   "establecimientos",
                "fuente":   "Cámara de Comercio (estimado)",
                "oficial":  False,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  250,
            },
            {
                "id":       "empleos_obra",
                "nombre":   "Empleos generados por obras",
                "unidad":   "empleos",
                "fuente":   "UP Secretarías",
                "oficial":  True,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  800,
            },
            {
                "id":       "inversion_privada",
                "nombre":   "Inversión privada atraída",
                "unidad":   "millones COP",
                "fuente":   "Cámara de Comercio (estimado)",
                "oficial":  False,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  15000,
            },
            {
                "id":       "emprendimientos",
                "nombre":   "Iniciativas de emprendimiento apoyadas",
                "unidad":   "unidades",
                "fuente":   "Secretaría de Desarrollo",
                "oficial":  False,
                "inverso":  False,
                "ref_min":  0,
                "ref_max":  80,
            },
        ],
    },
}

# ═══════════════════════════════════════════════════════════════════════════
# NORMALIZACIÓN
# ═══════════════════════════════════════════════════════════════════════════

def normalizar(valor, ref_min, ref_max, inverso=False):
    """
    Min-Max normalization → [0, 100].
    Si inverso=True, menor valor produce mayor puntaje (ej. criminalidad).
    """
    if ref_max == ref_min:
        return 0.0
    score = (valor - ref_min) / (ref_max - ref_min) * 100
    if inverso:
        score = 100 - score
    return max(0.0, min(100.0, score))


# ═══════════════════════════════════════════════════════════════════════════
# CÁLCULO PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════

def calcular_itt(valores_crudos: dict, periodo: str, meta: dict = None) -> dict:
    """
    Calcula el ITT a partir de los valores crudos de indicadores.

    Parámetros
    ----------
    valores_crudos : dict
        Estructura:  { "dim_id": { "ind_id": valor_numerico, ... }, ... }
        Ejemplo:
            {
                "seguridad": {
                    "homicidios": 28.4,
                    "hurtos": 312,
                    "percepcion_seguridad": 38,
                    "presencia_institucional": 2
                },
                ...
            }
    periodo : str
        Período de cálculo, ej. "2026-T1"
    meta : dict, optional
        Metadatos adicionales (version, cobertura_datos, etc.)

    Retorna
    -------
    dict  — Estructura compatible con data/itt_pulmon.json
    """
    resultado_dims = []
    itt_global = 0.0

    for dim_id, dim_cfg in DIMENSIONES.items():
        vals_dim = valores_crudos.get(dim_id, {})
        scores_ind = []
        indicadores_out = []

        for ind in dim_cfg["indicadores"]:
            valor = vals_dim.get(ind["id"])
            if valor is None:
                # Indicador sin dato: se omite del promedio dimensional
                indicadores_out.append({
                    "nombre":    ind["nombre"],
                    "valor":     None,
                    "unidad":    ind["unidad"],
                    "tendencia": "sin_dato",
                    "fuente":    ind["fuente"],
                    "oficial":   ind["oficial"],
                })
                continue

            score = normalizar(valor, ind["ref_min"], ind["ref_max"], ind["inverso"])
            scores_ind.append(score)
            indicadores_out.append({
                "nombre":    ind["nombre"],
                "valor":     valor,
                "unidad":    ind["unidad"],
                "tendencia": "alza",   # TODO: calcular vs periodo anterior
                "fuente":    ind["fuente"],
                "oficial":   ind["oficial"],
            })

        score_dim = sum(scores_ind) / len(scores_ind) if scores_ind else 0.0
        cobertura = round(len(scores_ind) / len(dim_cfg["indicadores"]) * 100)

        resultado_dims.append({
            "id":              dim_id,
            "nombre":          dim_cfg["nombre"],
            "icono":           dim_cfg["icono"],
            "peso":            dim_cfg["peso"],
            "score":           round(score_dim, 1),
            "score_anterior":  0.0,   # completar con periodo anterior
            "variacion":       0.0,   # completar comparando periodos
            "color":           dim_cfg["color"],
            "calidad_dato":    "oficial" if cobertura == 100 else "preliminar",
            "cobertura":       cobertura,
            "indicadores":     indicadores_out,
        })

        itt_global += score_dim * dim_cfg["peso"]

    # Cobertura global
    total_ind   = sum(len(d["indicadores"]) for d in DIMENSIONES.values())
    con_valor   = sum(
        sum(1 for ind in dim_cfg["indicadores"] if valores_crudos.get(dim_id, {}).get(ind["id"]) is not None)
        for dim_id, dim_cfg in DIMENSIONES.items()
    )
    cobertura_global = round(con_valor / total_ind * 100)
    fuentes_activas  = sum(
        1 for dim_id, dim_cfg in DIMENSIONES.items()
        for ind in dim_cfg["indicadores"]
        if valores_crudos.get(dim_id, {}).get(ind["id"]) is not None
    )

    resultado = {
        "meta": {
            "territorio":      "Pulmón de Oriente",
            "version":         meta.get("version", "preliminar") if meta else "preliminar",
            "periodo":         periodo,
            "fecha_calculo":   datetime.today().strftime("%Y-%m-%d"),
            "cobertura_datos": cobertura_global,
            "fuentes_activas": fuentes_activas,
            "fuentes_total":   total_ind,
            "nota":            meta.get("nota", "ITT calculado automáticamente.") if meta else "ITT calculado automáticamente.",
        },
        "itt_global": {
            "score":              round(itt_global, 1),
            "score_anterior":     0.0,
            "variacion":          0.0,
            "clasificacion":      clasificar_itt(itt_global),
            "rango":              rango_itt(itt_global),
            "periodo_comparacion": "N/A",
        },
        "dimensiones":   resultado_dims,
        "serie_temporal": [],   # completar acumulando periodos anteriores
        "barrios":        [],   # completar con datos por barrio
    }

    return resultado


def clasificar_itt(score: float) -> str:
    if score < 40:  return "Crítico"
    if score < 60:  return "Transformación Moderada"
    if score < 80:  return "Transformación Avanzada"
    return "Transformación Plena"


def rango_itt(score: float) -> list:
    if score < 40:  return [0, 40]
    if score < 60:  return [40, 60]
    if score < 80:  return [60, 80]
    return [80, 100]


# ═══════════════════════════════════════════════════════════════════════════
# LECTURA DESDE EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def leer_excel(ruta: str) -> dict:
    """
    Lee un Excel con la siguiente estructura esperada:
      Hoja "indicadores":
        Col A: dimension_id   (ej. "seguridad")
        Col B: indicador_id   (ej. "homicidios")
        Col C: valor          (número)

    Retorna dict { dim_id: { ind_id: valor } }
    """
    if not HAS_OPENPYXL:
        raise ImportError("Instala openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb["indicadores"] if "indicadores" in wb.sheetnames else wb.active

    valores = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        dim_id = str(row[0]).strip().lower().replace(" ", "_")
        ind_id = str(row[1]).strip().lower().replace(" ", "_")
        try:
            valor = float(row[2])
        except (TypeError, ValueError):
            continue
        valores.setdefault(dim_id, {})[ind_id] = valor

    wb.close()
    return valores


# ═══════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Calcula el ITT del Pulmón de Oriente.")
    parser.add_argument("--input",   default=None,        help="Ruta al Excel de indicadores (.xlsx)")
    parser.add_argument("--output",  default="data/itt_pulmon.json", help="Ruta del JSON de salida")
    parser.add_argument("--periodo", default=None,        help="Período (ej. 2026-T1). Default: actual.")
    parser.add_argument("--version", default="preliminar",help="'preliminar' u 'oficial'")
    args = parser.parse_args()

    # Período por defecto: trimestre actual
    if not args.periodo:
        hoy = datetime.today()
        trimestre = (hoy.month - 1) // 3 + 1
        args.periodo = f"{hoy.year}-T{trimestre}"

    # Leer datos
    if args.input:
        print(f"Leyendo indicadores desde: {args.input}")
        valores = leer_excel(args.input)
    else:
        print("No se especificó --input. Usando datos de ejemplo (mock).")
        valores = {
            "seguridad":         {"homicidios": 28.4, "hurtos": 312, "percepcion_seguridad": 38, "presencia_institucional": 2},
            "entorno_urbano":    {"espacio_publico": 4.2, "frentes_obra": 34, "luminarias": 210, "parques": 6},
            "movilidad":         {"tramos_viales": 3.8, "accesibilidad_mio": 72, "tiempo_desplazamiento": 38, "ciclorutas": 1.2},
            "desarrollo_social": {"ninos_programas": 1840, "cobertura_educativa": 88, "acceso_salud": 76, "familias_subsidio": 520},
            "actividad_economica":{"negocios_formalizados": 87, "empleos_obra": 340, "inversion_privada": 4200, "emprendimientos": 23},
        }

    meta = {"version": args.version, "nota": f"ITT {args.version} calculado para periodo {args.periodo}."}
    resultado = calcular_itt(valores, args.periodo, meta)

    # Guardar
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f"\n✓ ITT calculado: {resultado['itt_global']['score']}/100 — {resultado['itt_global']['clasificacion']}")
    print(f"✓ Cobertura de datos: {resultado['meta']['cobertura_datos']}%")
    print(f"✓ Archivo guardado: {output_path}")


if __name__ == "__main__":
    main()
